from openpyxl import load_workbook

# Cargo el archivo fuente
workbook_origen = load_workbook('afip.xlsx')
hoja_origen = workbook_origen.active

# Cargar el archivo destino
workbook_destino = load_workbook('client.xlsx')
hoja_destino = workbook_destino.active

# Copiar datos de una celda
dato = hoja_origen['A3'].value
hoja_destino['A2'] = dato

# Guardar el archivo modificado

workbook_destino.save('client_mod.xlsx')
